# coding=utf-8
import openpyxl


def compare_excel(path1, path2):
    wb_excel1 = openpyxl.load_workbook(filename=path1)
    wb_excel2 = openpyxl.load_workbook(filename=path2)

    sheet_names = wb_excel1.sheetnames

    error_messages = []
    sheet_index = 0
    for sheet1 in wb_excel1.worksheets:
        agency_name_index = 2
        drdpcompletion_name_index = 12
        sheet_name = sheet_names[sheet_index]
        # 跳过info的sheet
        if 'info' in sheet_name:
            sheet_index += 1
            continue
        sheet2 = wb_excel2[sheet_name]
        row_index = 1
        for row1 in sheet1:
            cell_index = 1
            for cell1 in row1:
                cell2 = sheet2.cell(row=row_index, column=cell_index)
                # 获取agency列索引
                if row_index == 1 and cell1.value == 'agency' and cell2.value == 'agency':
                    agency_name_index = cell_index
                # 如果是agency列，则跳过不比较
                if cell_index == agency_name_index:
                    cell_index += 1
                    continue
                # 获取drdpcompletion列索引
                if row_index == 1 and cell1.value == 'drdpcompletion' and cell2.value == 'drdpcompletion':
                    drdpcompletion_name_index = cell_index
                # 如果是drdpcompletion列，则跳过不比较
                if cell_index == drdpcompletion_name_index:
                    cell_index += 1
                    continue

                if cell1.value == 'ignore' or cell2.value == 'ignore':
                    cell_index += 1
                    continue

                if (cell1.value is None and cell2.value == '') or (cell1.value == '' and cell2.value is None):
                    cell_index += 1
                    continue

                if cell1.value != cell2.value:
                    error_messages.append('' + sheet_name + ' row: ' + str(row_index) + ' col: ' + str(cell_index) + ' ( ' + str(cell1.value) + ' != ' + str(cell2.value) + ' ) ')
                cell_index += 1
            row_index += 1
        sheet_index += 1
    if len(error_messages) > 0:
        raise ValueError(error_messages)
    print 'Content is equal!'

compare_excel("E:\\testFile\DRDP2015_IT_sample.xlsx","E:\\testFile\DRDPTECH_Site01_Class01_DRDP2015_INFANT_TODDLER_05_02_2018_14%3A25%3A37.xlsx")
